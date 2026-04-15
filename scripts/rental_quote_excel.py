#!/usr/bin/env python3
"""
Generador de cotizaciones Excel — SP Rental (línea 4000)

Uso:
    python rental_quote_excel.py datos.json
    python rental_quote_excel.py datos.json -o cotizacion.xlsx
    cat datos.json | python rental_quote_excel.py

Estructura del JSON de entrada:
{
  "quote":  { ...campos de Quote },
  "items":  [ ...lista de QuoteItem ],
  "faena":  { ...RentalFaena } | null
}

El campo "quote_number" ya trae el correlativo completo (ej. "4000-0036").
"""

import sys
import json
import argparse
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta ───────────────────────────────────────────────────────────────────

BLACK  = "FF000000"
WHITE  = "FFFFFFFF"
YELLOW = "FFFFD700"
GRAY1  = "FF333333"
GRAY2  = "FF666666"
GRAY5  = "FFF5F5F5"
BORDER = "FFDDDDDD"

# ── Helpers de estilo ─────────────────────────────────────────────────────────

def font(bold=False, size=9, color=BLACK, italic=False, name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_all(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom(style="thin"):
    s = Side(style=style)
    return Border(bottom=s)

THIN  = border_all("thin")
LIGHT = border_all()          # thin también; placeholder para legibilidad

# ── Layout de columnas ────────────────────────────────────────────────────────
#
#  A      B         C       D    E      F           G
#  ITEM   DESCRIP.  DISP.   UM   CANT.  P.UNIT      TOTAL
#  5      38        14      8    8      15          15
#
COLS = {
    "A": 5,
    "B": 38,
    "C": 14,
    "D": 8,
    "E": 8,
    "F": 15,
    "G": 15,
}
N = 7   # total columnas (A=1 … G=7)


# ─────────────────────────────────────────────────────────────────────────────


class RentalQuoteExcel:

    def __init__(self, quote: dict, items: list, faena: dict | None):
        self.quote = quote
        self.items = items
        self.faena = faena or {}

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = f"Cotización {quote.get('quote_number', '')}"
        self.ws.sheet_view.showGridLines = False

        for col, w in COLS.items():
            self.ws.column_dimensions[col].width = w

        self._row = 1   # cursor de fila

    # ── Primitivas ────────────────────────────────────────────────────────────

    def _c(self, col: int, value=None):
        """Devuelve la celda en la fila actual, columna col (1-based)."""
        return self.ws.cell(row=self._row, column=col, value=value)

    def _merge(self, c1: int, c2: int, r_offset: int = 0):
        """Fusiona columnas c1..c2 en la fila actual + r_offset."""
        r = self._row + r_offset
        self.ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

    def _merge_rows(self, r1: int, c1: int, r2: int, c2: int):
        """Fusiona bloque de celdas (filas y columnas absolutas)."""
        self.ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def _next(self, height: int | None = None):
        if height:
            self.ws.row_dimensions[self._row].height = height
        self._row += 1

    def _blank(self, height: int = 8):
        self.ws.row_dimensions[self._row].height = height
        self._row += 1

    # ── Secciones ─────────────────────────────────────────────────────────────

    def _build_header(self):
        """Encabezado: título + datos de SP RENTAL."""

        # ── Fila 1: OFERTA DE ARRIENDO ──
        self._merge(1, N)
        c = self._c(1, "OFERTA DE ARRIENDO")
        c.font      = font(bold=True, size=14)
        c.alignment = align("center", "center")
        self._next(22)

        # ── Fila 2: disclaimer ──
        self._merge(1, N)
        c = self._c(1, "Esta cotización refleja oferta al momento de su emisión y puede variar en el tiempo.")
        c.font      = font(italic=True, size=7, color=GRAY2)
        c.alignment = align("center", "center")
        self._next(12)

        # ── Filas 3-5: Logo (cols A-B) + datos empresa (cols C-G) ──
        logo_row = self._row

        # Logo (3 filas fusionadas)
        self._merge_rows(logo_row, 1, logo_row + 2, 2)
        c = self.ws.cell(row=logo_row, column=1, value="SP\nRENTAL")
        c.font      = font(bold=True, size=18)
        c.alignment = align("center", "center", wrap=True)
        c.border    = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"),  bottom=Side(style="medium"),
        )

        # Empresa (derecha)
        empresa_data = [
            ("SP RENTAL", True, 10),
            ("RUT: 77.437.583-k  |  Campos de deportes 780 – Ñuñoa  |  +56 9 39107419", False, 8),
            ("RL: SERGIO ASPE, PAOLO ALARCÓN", False, 8),
        ]
        for offset, (text, bold_, size_) in enumerate(empresa_data):
            self._merge_rows(logo_row + offset, 3, logo_row + offset, N)
            c = self.ws.cell(row=logo_row + offset, column=3, value=text)
            c.font      = font(bold=bold_, size=size_)
            c.alignment = align("right", "center")
            self.ws.row_dimensions[logo_row + offset].height = 14

        self._row = logo_row + 3

    def _build_client(self):
        """Datos del cliente + número de cotización + fila SR."""
        q       = self.quote
        company = q.get("company") or {}
        contact = q.get("contact") or {}

        # Formatear fecha
        raw_date = q.get("quote_date", "")
        try:
            quote_date = datetime.fromisoformat(raw_date).strftime("%d-%m-%Y")
        except Exception:
            quote_date = raw_date

        quote_num = q.get("quote_number", "")

        fields = [
            ("Empresa:",   company.get("name", "")),
            ("RUT:",       company.get("rut", "")),
            ("Email:",     contact.get("email", "")),
            ("Dirección:", company.get("address", "")),
            ("Teléfono:",  contact.get("phone", "")),
        ]

        for i, (label, value) in enumerate(fields):
            row = self._row + i

            # Label
            c = self.ws.cell(row=row, column=1, value=label)
            c.font      = font(bold=True, size=8.5)
            c.alignment = align("left", "center")

            # Valor (cols B-D)
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            c = self.ws.cell(row=row, column=2, value=value)
            c.font      = font(size=8.5)
            c.alignment = align("left", "center")

            # Número de cotización (solo primeras 2 filas)
            if i == 0:
                self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=N)
                c = self.ws.cell(row=row, column=5, value=f"COTIZACIÓN:   {quote_num}")
                c.font      = font(bold=True, size=11)
                c.alignment = align("right", "center")

            elif i == 1:
                self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=N)
                c = self.ws.cell(row=row, column=5, value=f"FECHA:   {quote_date}")
                c.font      = font(bold=True, size=9)
                c.alignment = align("right", "center")

            self.ws.row_dimensions[row].height = 14

        self._row += len(fields) + 1

        # ── Fila SR (fondo amarillo) ──
        self._merge(1, N)
        contact_name = (contact.get("full_name") or "").upper()
        c = self._c(1, f"SR.     {contact_name}")
        c.font      = font(bold=True, size=9)
        c.fill      = fill(YELLOW)
        c.alignment = align("left", "center")
        self._next(18)

    def _build_table_uf(self):
        """Tabla de ítems en UF + bloque de totales UF."""
        uf_items = [i for i in self.items if i.get("currency") == "UF"]
        if not uf_items:
            return

        headers = ["ITEM", "DESCRIPCIÓN", "DISP.", "UM", "CANT.", "P. UNIT\n(UF)", "TOTAL\n(UF)"]
        self._write_table_header(headers)

        for idx, item in enumerate(uf_items):
            bg = GRAY5 if idx % 2 == 0 else WHITE
            self._write_item_row(idx + 1, item, bg, currency="UF")

        self._write_totals_uf()

    def _build_table_clp(self):
        """Tabla de ítems en CLP + bloque de totales CLP + NOTAS."""
        clp_items = [i for i in self.items if i.get("currency") != "UF"]

        headers = ["ITEM", "DESCRIPCIÓN", "DISP.", "UM", "CANT.", "P. UNIT\n(CLP)", "TOTAL\n(CLP)"]
        self._write_table_header(headers)

        for idx, item in enumerate(clp_items):
            bg = GRAY5 if idx % 2 == 0 else WHITE
            self._write_item_row(idx + 1, item, bg, currency="CLP")

        self._write_totals_clp()

    def _write_table_header(self, headers: list[str]):
        for col, text in enumerate(headers, 1):
            c = self._c(col, text)
            c.font      = font(bold=True, size=7.5, color=WHITE)
            c.fill      = fill(BLACK)
            c.alignment = align("center", "center", wrap=True)
            c.border    = THIN
        self._next(24)

    def _write_item_row(self, idx: int, item: dict, bg: str, currency: str):
        num_fmt = '#,##0.00' if currency == "UF" else '$#,##0'

        vals = [
            (idx,                               "center", None),
            (item.get("description", ""),       "left",   None),
            (item.get("availability", ""),      "center", None),
            (item.get("unit", ""),              "center", None),
            (item.get("quantity", 0),           "center", '#,##0'),
            (item.get("unit_price", 0),         "right",  num_fmt),
            (item.get("total_price", 0),        "right",  num_fmt),
        ]

        for col, (val, h_align, fmt) in enumerate(vals, 1):
            c = self._c(col, val)
            c.font      = font(size=8)
            c.fill      = fill(bg)
            c.border    = THIN
            c.alignment = align(h_align, "center", wrap=(col == 2))
            if fmt:
                c.number_format = fmt

        self._next(14)

    def _write_totals_uf(self):
        """4 filas de totales: Sub Total / Descuento / IVA / Total UF."""
        subtotal = self.quote.get("subtotal_uf", 0)
        disc_pct = self.quote.get("discount_pct_uf", 0)
        iva      = self.quote.get("iva_uf", 0)
        total    = self.quote.get("total_uf", 0)
        disc_val = -(subtotal * disc_pct / 100) if disc_pct else None

        rows = [
            ("Sub Total",          subtotal, False),
            (f"Descuento {int(disc_pct) if disc_pct == int(disc_pct) else disc_pct}%",
                                   disc_val, False),
            ("IVA (19%)",          iva,      False),
            ("Total UF",           total,    True),
        ]
        for label, value, bold_ in rows:
            # Cols A-E vacías (pero con borde limpio en la izquierda de la tabla)
            self._merge(1, 5)
            c_left = self._c(1)
            c_left.border = Border()

            # Col F: etiqueta
            c = self.ws.cell(row=self._row, column=6, value=label)
            c.font      = font(bold=bold_, size=8)
            c.alignment = align("right", "center")
            c.border    = THIN

            # Col G: valor
            c = self.ws.cell(row=self._row, column=7, value=value)
            c.font         = font(bold=bold_, size=8)
            c.alignment    = align("right", "center")
            c.number_format = '#,##0.00'
            c.border       = THIN
            if bold_:
                c.fill = fill(GRAY5)

            self._next(14)

    def _write_totals_clp(self):
        """
        NOTAS (cols A-D, filas r…r+3) y totales CLP (cols E-G) lado a lado.
        """
        subtotal = self.quote.get("subtotal_clp", 0)
        disc_pct = self.quote.get("discount_pct", 0)
        iva      = self.quote.get("iva_clp", 0)
        total    = self.quote.get("total_clp", 0)
        disc_val = -(subtotal * disc_pct / 100) if disc_pct else None
        notes    = self.quote.get("notes", "") or ""

        totals_rows = [
            ("Sub Total",          subtotal, False),
            (f"Descuento {int(disc_pct) if disc_pct == int(disc_pct) else disc_pct}%",
                                   disc_val, False),
            ("IVA (19%)",          iva,      False),
            ("Total CLP",          total,    True),
        ]

        start = self._row

        # ── NOTAS (bloque fusionado A-D, 4 filas) ──
        # Etiqueta "NOTAS:" en A del primer renglón
        c = self.ws.cell(row=start, column=1, value="NOTAS:")
        c.font      = font(bold=True, size=7.5)
        c.alignment = align("left", "top")

        # Texto de notas: cols B-D, 4 filas
        self._merge_rows(start, 2, start + 3, 4)
        c = self.ws.cell(row=start, column=2, value=notes)
        c.font      = font(italic=True, size=7.5, color=GRAY2)
        c.alignment = align("left", "top", wrap=True)
        c.border    = border_all("thin")

        # ── Totales CLP (cols E-G) ──
        for offset, (label, value, bold_) in enumerate(totals_rows):
            row = start + offset

            # Col E: vacía (separador)
            self.ws.cell(row=row, column=5).border = Border()

            # Col F: etiqueta
            c = self.ws.cell(row=row, column=6, value=label)
            c.font      = font(bold=bold_, size=8)
            c.alignment = align("right", "center")
            c.border    = THIN

            # Col G: valor
            c = self.ws.cell(row=row, column=7, value=value)
            c.font         = font(bold=bold_, size=8)
            c.alignment    = align("right", "center")
            c.number_format = '$#,##0'
            c.border       = THIN
            if bold_:
                c.fill = fill(GRAY5)

            self.ws.row_dimensions[row].height = 14

        self._row = start + len(totals_rows)

    def _build_faena(self):
        """REQUERIMIENTOS DE FAENA."""
        f = self.faena

        self._blank(6)

        # Título sección
        self._merge(1, N)
        c = self._c(1, "REQUERIMIENTOS DE FAENA")
        c.font      = font(bold=True, size=9)
        c.alignment = align("left", "center")
        self._next(16)

        subsections = [
            ("1. Uso de la Maquinaria", [
                ("Lugar",                                   f.get("lugar_faena", "")),
                ("Uso específico que se le dará en faena",  f.get("uso_especifico", "")),
                ("Plazos de ejecución de proyecto o contrato", f.get("plazo_ejecucion", "")),
            ]),
            ("2. Equipamiento de Máquinas", [
                ("Estándar minero requerido para ingreso a faena", f.get("estandar_minero", "")),
                ("Equipamiento adicional",                         f.get("equipamiento_add", "")),
            ]),
            ("3. Mantenciones Correctivas", [
                ("Responsable",                      f.get("responsable_mant", "")),
                ("Se realiza en faena",               f.get("realiza_en_faena", "")),
                ("Kit de repuestos",                  f.get("kit_repuestos", "")),
                ("Plazos de retiro ante desperfecto", f.get("plazo_retiro", "")),
            ]),
        ]

        for sub_title, sub_rows in subsections:
            # Sub-título
            self._merge(1, N)
            c = self._c(1, sub_title)
            c.font      = font(bold=True, size=8.5)
            c.alignment = align("left", "center")
            self._next(14)

            for label, value in sub_rows:
                # Cols A-B: etiqueta (fondo gris)
                self._merge(1, 2)
                c = self._c(1, label)
                c.font      = font(bold=True, size=8)
                c.fill      = fill(GRAY5)
                c.border    = THIN
                c.alignment = align("left", "center", wrap=True)

                # Cols C-G: valor
                self._merge(3, N)
                c = self.ws.cell(row=self._row, column=3,
                                 value=value if value else "")
                c.font      = font(italic=True, size=8, color=GRAY2)
                c.border    = THIN
                c.alignment = align("left", "center", wrap=True)

                self._next(16)

            self._blank(4)

    def _build_footer(self):
        """Datos bancarios, alcances y disclaimer final."""
        q = self.quote

        self._blank(6)

        # ── Datos bancarios ──
        self._merge(1, N)
        c = self._c(1, "DATOS BANCARIOS")
        c.font      = font(bold=True, size=8)
        c.alignment = align("left", "center")
        self._next(14)

        bank = [
            ("SYP INGENIERÍA Y CONSTRUCCIÓN SPA", "Banco de Chile"),
            ("77.437.583-K",                      "Cuenta Corriente"),
            ("",                                  "1450808504"),
        ]
        for left, right in bank:
            self._merge(1, 3)
            c = self._c(1, left)
            c.font      = font(size=8)
            c.alignment = align("left", "center")

            self._merge(4, N)
            c = self.ws.cell(row=self._row, column=4, value=right)
            c.font      = font(size=8)
            c.alignment = align("left", "center")
            self._next(13)

        self._blank(6)

        # ── Alcances ──
        # Cabecera negra
        self._merge(1, N)
        c = self._c(1, "ALCANCES")
        c.font      = font(bold=True, size=8, color=WHITE)
        c.fill      = fill(BLACK)
        c.alignment = align("left", "center")
        self._next(14)

        alcances = [
            ("CONDICIONES COMERCIALES", q.get("payment_conditions", "") or ""),
            ("PLAZO DE ENTREGA",        q.get("delivery_time", "")      or ""),
            ("GARANTÍA",               q.get("warranty", "")           or ""),
            ("DESPACHO",               q.get("dispatch", "")           or ""),
            ("CONTACTO",               q.get("contact_scope", "")      or ""),
        ]
        for label, value in alcances:
            # Label (gris)
            self._merge(1, 3)
            c = self._c(1, label)
            c.font      = font(bold=True, size=8)
            c.fill      = fill(GRAY5)
            c.border    = border_all("thin")
            c.alignment = align("left", "center")

            # Valor
            self._merge(4, N)
            c = self.ws.cell(row=self._row, column=4, value=value)
            c.font      = font(size=8)
            c.border    = border_all("thin")
            c.alignment = align("left", "center")
            self._next(14)

        self._blank(12)

        # ── Disclaimer final ──
        self._merge(1, N)
        c = self._c(1, "Esta cotización refleja la disponibilidad al momento de su emisión y puede variar en el tiempo.")
        c.font      = font(italic=True, size=7, color=GRAY2)
        c.alignment = align("center", "center")
        self._next(12)

    # ── Punto de entrada ──────────────────────────────────────────────────────

    def build(self, output_path: str) -> str:
        self._build_header()
        self._blank(8)
        self._build_client()
        self._blank(8)

        has_uf = self.quote.get("has_uf_section", False)
        if has_uf:
            self._build_table_uf()
            self._blank(10)

        self._build_table_clp()
        self._build_faena()
        self._build_footer()

        self.wb.save(output_path)
        return output_path


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Genera cotización Excel para SP Rental (línea 4000)"
    )
    parser.add_argument(
        "input", nargs="?",
        help="Archivo JSON con datos (omitir para leer de stdin)"
    )
    parser.add_argument(
        "-o", "--output",
        help="Ruta del archivo .xlsx generado (por defecto: cotizacion-<numero>.xlsx)"
    )
    args = parser.parse_args()

    if args.input:
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = json.load(sys.stdin)

    quote = data.get("quote", {})
    items = data.get("items", [])
    faena = data.get("faena", None)

    if quote.get("business_line") != 4000:
        print("Error: este script es exclusivo para cotizaciones Rental (business_line = 4000).", file=sys.stderr)
        sys.exit(1)

    output = args.output or f"cotizacion-{quote.get('quote_number', 'DRAFT')}.xlsx"

    gen = RentalQuoteExcel(quote, items, faena)
    gen.build(output)
    print(f"Cotización generada: {output}")


if __name__ == "__main__":
    main()
